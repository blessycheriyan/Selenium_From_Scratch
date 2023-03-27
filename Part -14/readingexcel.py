
import openpyxl

# file -> Workbook -> Sheets -> Rows -> cells

file= 'D:\Selenium\student_marks.xlsx'
workbook = openpyxl.load_workbook(file)
sheet = workbook['student_marks']

rows = sheet.max_row # count number of rows in a excel sheet
cols = sheet.max_column # count number of column in a excel sheet

# Reading all the rows in a excel sheet
for r in range(1,rows +1):
    for c in range(1,cols +1):
        print(sheet.cell(r,c).value, end=' ')
    print()    