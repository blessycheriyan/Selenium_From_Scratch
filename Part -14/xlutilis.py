import openpyxl
from openpyxl.styles import Color, PatternFill

def getrowcount(file,student_marks):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook['student_marks']
    return (sheet.max_row)

def getColumncount(file,student_marks):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook['student_marks']
    return sheet.max_column

def readdata(file,student_marks,rowno,columno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook['student_marks']
    return  print(sheet.cell(rowno,columno).value, end=' ')

def writrdata(file,student_marks,rowno,columno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook['student_marks']
    sheet.cell(rowno,columno).value = data    
    workbook.save(file)  

def filegreencolor(file,student_marks,rowno,columno) :
     workbook = openpyxl.load_workbook(file)
     sheet = workbook['student_marks']
     greenFill = PatternFill(start_color = '60b212',
                         end_color = '60b212' ,
                         fill_type ='solid'
                         )
     sheet.cell(rowno,columno).fill =greenFill
     workbook.save(file)  

def fileredcolor(file,student_marks,rowno,columno) :
     workbook = openpyxl.load_workbook(file)
     sheet = workbook['student_marks']
     redFill = PatternFill(start_color = 'ff0000',
                         end_color = 'ff0000' ,
                         fill_type ='solid'
                         )
     sheet.cell(rowno,columno).fill =redFill
     workbook.save(file)      