import openpyxl

# file -> Workbook -> Sheets -> Rows -> cells

# file= 'D:\\Selenium\\test_data.xlsx'
# workbook = openpyxl.load_workbook(file)

# sheet = workbook['Sheet1']
# # sheet = workbook.active

# for r in range(1,6):
#     for c in range(1,4):
#          sheet.cell(r,c).value ='Welcomes'
# workbook.save(file)       

# Multiple data files
file= 'D:\\Selenium\\test_one.xlsx'
workbook = openpyxl.load_workbook(file)

sheet = workbook['Sheet1']

sheet.cell(1,1).value =123
sheet.cell(1,2).value ='James'
sheet.cell(1,3).value ='Software Developer'

sheet.cell(2,1).value =124
sheet.cell(2,2).value ='John'
sheet.cell(2,3).value ='Project Manager'

sheet.cell(3,1).value =125
sheet.cell(3,2).value ='David'
sheet.cell(3,3).value ='Team Lead'


workbook.save(file)    
  